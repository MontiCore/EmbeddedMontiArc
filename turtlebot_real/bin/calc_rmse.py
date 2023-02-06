from math import sqrt
from scipy.interpolate import interp1d
import numpy as np
import openpyxl
import sys

def import_from_excel(file_path):
    """
    This function takes in a file path as an argument and returns the x and y lists
    that were exported using the export_to_excel function.
    """
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)
    
    # Get the first sheet
    sheet = wb.active
    
    # Extract the x and y lists from the sheet
    x = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row]
    y = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2) for cell in row]
    
    return x[1:], y[1:]

def calculate_rmse(recorded_x, recorded_y, actual_x, actual_y):
    """
    This function takes four lists of x and y coordinates as input, and calculates the root mean square error between them.
    recorded_x: list of recorded x coordinates
    recorded_y: list of recorded y coordinates
    actual_x: list of actual x coordinates
    actual_y: list of actual y coordinates
    """
    rmse = 0
    for i in range(len(recorded_x)):
        rmse += ((recorded_x[i] - actual_x[i]) ** 2 + (recorded_y[i] - actual_y[i]) ** 2)
    rmse = sqrt(rmse / len(recorded_x))
    return rmse


def make_lists_same(recorded_x, recorded_y, actual_x, actual_y, num_points):
    """
    This function takes four lists and one integer as input, the four lists are the recorded x coordinates, recorded y coordinates, actual x coordinates, and actual y coordinates. 
    The integer is the number of points which we want the lists to have.
    recorded_x: list of recorded x coordinates
    recorded_y: list of recorded y coordinates
    actual_x: list of actual x coordinates
    actual_y: list of actual y coordinates
    num_points: number of points in the final lists
    """
    f_recorded_x = interp1d(np.arange(len(recorded_x)), recorded_x)
    f_recorded_y = interp1d(np.arange(len(recorded_y)), recorded_y)
    f_actual_x = interp1d(np.arange(len(actual_x)), actual_x)
    f_actual_y = interp1d(np.arange(len(actual_y)), actual_y)
    resampled_recorded_x = f_recorded_x(np.linspace(0, len(recorded_x)-1, num_points))
    resampled_recorded_y = f_recorded_y(np.linspace(0, len(recorded_y)-1, num_points))
    resampled_actual_x = f_actual_x(np.linspace(0, len(actual_x)-1, num_points))
    resampled_actual_y = f_actual_y(np.linspace(0, len(actual_y)-1, num_points))
    return resampled_recorded_x, resampled_recorded_y, resampled_actual_x, resampled_actual_y



if __name__ == "__main__":
    
    try:
        exp_name = sys.argv[1]
    except IndexError:
        print("No argument for the experiment is provided")
        sys.exit(1)
    
    actual_x, actual_y = import_from_excel('../excel_traj_files/act_traj_' + exp_name +'.xlsx')
    recorded_x, recorded_y = import_from_excel('../excel_traj_files/rec_traj_' + exp_name +'.xlsx')
    
    num_points = min(len(actual_x), len(recorded_x))
    
    resampled_recorded_x, resampled_recorded_y, resampled_actual_x, resampled_actual_y = make_lists_same(recorded_x, recorded_y, actual_x, actual_y, num_points)
    rmse = calculate_rmse(resampled_recorded_x, resampled_recorded_y, resampled_actual_x, resampled_actual_y)
    
    if rmse >= 1:
        raise ValueError("Calculated RMSE is too high: {:.2f}".format(rmse))
    elif rmse < 1:
        print("Calculated RMSE is low: {:.2f}".format(rmse))
